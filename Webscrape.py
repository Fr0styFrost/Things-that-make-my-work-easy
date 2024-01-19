import requests
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

# Set the path to your ChromeDriver executable
chromedriver_path = r'C:\Users\Frost\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe'

## Create ChromeOptions object
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument(f"executable_path={chromedriver_path}")

# Create a new instance of the Chrome driver with options
driver = webdriver.Chrome(options=chrome_options)

# Replace the URL with the one you want to scrape
search = 'people who work as Controls System engineer in UK'
url = 'https://www.google.com/search?q=' + search + ' linkedin.com/in/'
workbookname = search +'.xlsx'
driver.get(url)

# Sleep to allow time for the cookies banner to appear (adjust as needed)-
time.sleep(2)

# Find and click the "Accept" button for cookies (replace with actual HTML/CSS selectors)
accept_button = driver.find_element(By.XPATH, "//button[@id='L2AGLb' and contains(@class, 'tHlp8d')]//div[text()='Accept all']")
accept_button.click()
print ('om nom nom... cookies')
time.sleep(5)


# Function to check for new data and append to the sheet
def check_and_append_data():
    # Parse the HTML content using BeautifulSoup
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')

    # Example: Extracting all the links from the webpage
    links = [link.get('href') for link in soup.find_all('a')]

    # Load existing data from the sheet
    # If the file exists, read the existing data
    existing_data = [list(row) for row in existing_ws.iter_rows(min_row=2, values_only=True)]

    # Check for duplicates and append new data
    specific_domain = 'https://uk.linkedin.com/in/'
        
    for link in links:
        if link and link.startswith(specific_domain) and link not in (existing_row[0] for existing_row in existing_data):
            display_text = soup.find('a', href=link).get_text()
            existing_ws.append([link, display_text])  # Add new data to the sheet

        # Save the workbook
    existing_wb.save(workbookname)
try:
        # Load existing data from the sheet
        existing_wb = load_workbook(workbookname)
        existing_ws = existing_wb.active
        print ('File Opened')
except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        existing_wb = Workbook()
        existing_ws = existing_wb.active
        existing_ws.append(['Link', 'Display Text'])  # Add headers    
        print ('File Created')
        
for i in range(25):
    try:
        # Check if the button is present
        more_results_button = driver.find_element(By.XPATH, "//div[@class='GNJvt ipz2Oe']//span[@class='RVQdVd' and text()='More results']")
        
        # If the button is present, click it
        more_results_button.click()
        
        # Sleep to allow time for the new content to load (adjust as needed)
        time.sleep(2)
        print ('clicked it')
    except:
        pass
    # Scroll down using WebDriver
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    print ('they see me scrolling')
    # Sleep to allow time for the page to scroll and new content to load (adjust as needed)
    time.sleep(2)
    print ('they waiting')
    time.sleep(3)
    # Re-run the check for new data and append to the sheet
    check_and_append_data()
    print ('Patrollin and tryna catch me Scraping data')
    print (i)

# Close the browser window

print("Scraped data has been stored in "+ workbookname)
driver.quit()
# Close the browser window
#driver.quit()